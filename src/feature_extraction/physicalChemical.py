import openpyxl


class PhysicalChemicalType(object):
    DiDNA_original = r"./physicalchemical/original/DiDNA-original.xlsx"
    DiRNA_original = r"./physicalchemical/original/DiRNA-original.xlsx"
    MonoDNA_original = r"./physicalchemical/original/MonoDNA-original.xlsx"
    TriDNA_original = r"./physicalchemical/original/TriDNA-original.xlsx"

    DiDNA_standardized = r"./physicalchemical/standard/DiDNA-standardized.xlsx"
    DiRNA_standardized = r"./physicalchemical/standard/DiRNA-standardized.xlsx"
    MonoDNA_standardized = r"./physicalchemical/standard/MonoDNA-standardized.xlsx"
    TriDNA_standardized = r"./physicalchemical/standard/TriDNA-standardized.xlsx"


class PhysicalChemical:
    """
    PhysicalChemical values
    save format:PC[physical_chemical][base_pair]
    """

    def __init__(self, physical_chemical_type):
        self.pc_dict = self.get_pc_dict(physical_chemical_type)

    def get_pc_dict(self, xlsl):
        # PhysicalChemical()
        wb = openpyxl.load_workbook(xlsl)
        ws = wb.active  # 当前活跃的表单

        max_row = ws.max_row
        max_col = ws.max_column

        pc_name_list = [cell.value for cell in ws['B'][1:]]
        # print(pc_name_list)
        base_name_list = [str(cell.value).split(" ")[0] for cell in ws[1][4:max_col]]
        # print(base_name_list)

        pc_dict = {}
        pc_row = ws[1:max_row][1:]
        for row, pc_name in zip(pc_row, pc_name_list):
            pc_row_dict = {}
            for cell, base_name in zip(row[4:], base_name_list):
                pc_row_dict.update({base_name: cell.value})
            pc_dict.update({pc_name: pc_row_dict})
            # for cell_col, base in zip(ws.columns):
            #     pc_row_dict
        return pc_dict


if __name__ == '__main__':
    pc_dict = PhysicalChemical(PhysicalChemicalType.DiDNA_standardized).pc_dict
    print(pc_dict)
    print(pc_dict['Propeller twist'])
    print(pc_dict['Propeller twist']['AT'])
